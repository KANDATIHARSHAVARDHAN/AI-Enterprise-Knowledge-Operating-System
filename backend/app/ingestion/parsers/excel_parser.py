"""
EKOS Excel Parser
Extracts data from .xlsx files using openpyxl, converts to structured text.
"""

from pathlib import Path
import openpyxl
from app.utils.logger import logger
from app.utils.exceptions import IngestionError


class ExcelParser:
    """Parse Excel (.xlsx) files and extract data as structured text."""

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}

    def parse(self, file_path: str) -> list[dict]:
        """
        Parse an Excel file and extract data per sheet.

        Returns:
            List of dicts, one per sheet
        """
        path = Path(file_path)
        if not path.exists():
            raise IngestionError(f"File not found: {file_path}", filename=path.name)

        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)
            pages = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                if not rows:
                    continue

                # First row as header
                headers = [str(cell or f"Col{i}") for i, cell in enumerate(rows[0], 1)]

                # Build text representation
                lines = [f"Sheet: {sheet_name}"]
                lines.append(" | ".join(headers))
                lines.append("-" * 60)

                for row in rows[1:]:
                    cells = [str(cell or "") for cell in row]
                    # Skip completely empty rows
                    if any(c.strip() for c in cells):
                        lines.append(" | ".join(cells))

                content = "\n".join(lines)

                pages.append({
                    "content": content,
                    "metadata": {
                        "source": path.name,
                        "file_type": "xlsx",
                        "sheet_name": sheet_name,
                        "row_count": len(rows) - 1,
                        "column_count": len(headers),
                        "columns": headers,
                        "file_path": str(path),
                    }
                })

            wb.close()
            logger.info(f"Parsed Excel: {path.name} ({len(pages)} sheets)")
            return pages

        except Exception as e:
            raise IngestionError(f"Failed to parse Excel: {e}", filename=path.name)
